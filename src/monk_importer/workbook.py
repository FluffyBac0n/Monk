from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .geometry import (
    PolylineMatch,
    closest_polyline_match,
    cumulative_distances_km,
    point_distances_to_polyline_km,
)
from .models import (
    Detour,
    DetourConnection,
    Excursion,
    GeoPointValue,
    ImportPayload,
    Lodging,
    LodgingContact,
    RouteChunk,
    RouteMarker,
    RouteMetadata,
    Stage,
    StageServices,
)
from .utils import (
    clean_text,
    format_time,
    lodging_id,
    parse_coordinates,
    parse_price,
    slugify,
    stage_id,
    to_bool_yn,
    to_float,
    to_int,
)


STAGES_SHEET = "E4 Cyprus STAGES"
LODGING_SHEET = "LODGING"
ROUTE_SHEET = "Map & Elevation"
EXCURSIONS_SHEET = "Excursions"
DETOURS_SHEET = "Detours"


def build_import_payload(
    workbook_path: Path,
    *,
    trail_id: str = "cyprus-e4",
    route_chunk_size: int = 500,
    route_version: int = 1,
) -> ImportPayload:
    workbook_values = load_workbook(workbook_path, data_only=True, read_only=True)

    stages = parse_stages(workbook_values[STAGES_SHEET])
    stage_lookup = {stage.name: stage.id for stage in stages}
    stage_slug_lookup = {slugify(stage.name): stage.id for stage in stages}
    lodgings, lodging_warnings = parse_lodgings(
        workbook_values[LODGING_SHEET],
        stage_lookup,
        stage_slug_lookup,
    )
    route_metadata, route_chunks, route_markers = parse_route(
        workbook_values[ROUTE_SHEET],
        stages_by_name=stage_lookup,
        stages_by_slug=stage_slug_lookup,
        chunk_size=route_chunk_size,
        version=route_version,
    )
    excursions: list[Excursion] = []
    excursion_route_chunks: dict[str, list[RouteChunk]] = {}
    excursion_warnings: list[str] = []
    if EXCURSIONS_SHEET in workbook_values.sheetnames:
        excursions, excursion_route_chunks, excursion_warnings = parse_excursions(
            workbook_values[EXCURSIONS_SHEET],
            stages_by_sequence={stage.sequence: stage for stage in stages},
            main_route_points=flatten_route_points(route_chunks),
            chunk_size=route_chunk_size,
            version=route_version,
        )
    detours: list[Detour] = []
    detour_route_chunks: dict[str, list[RouteChunk]] = {}
    detour_warnings: list[str] = []
    if DETOURS_SHEET in workbook_values.sheetnames:
        detours, detour_route_chunks, detour_warnings = parse_detours(
            workbook_values[DETOURS_SHEET],
            stages=stages,
            main_route_points=flatten_route_points(route_chunks),
            chunk_size=route_chunk_size,
            version=route_version,
        )

    trail = {
        "id": trail_id,
        "name": "Cyprus E4",
        "country": "Cyprus",
        "stageCount": len(stages),
        "lodgingCount": len(lodgings),
        "excursionCount": len(excursions),
        "detourCount": len(detours),
        "routePointCount": route_metadata.pointCount,
        "totalDistanceKm": route_metadata.totalDistanceKm,
        "startStageId": stages[0].id if stages else None,
        "startStageName": stages[0].name if stages else None,
        "endStageId": stages[-1].id if stages else None,
        "endStageName": stages[-1].name if stages else None,
        "routeVersion": route_version,
    }

    return ImportPayload(
        trailId=trail_id,
        trail=trail,
        stages=stages,
        lodgings=lodgings,
        routeMetadata=route_metadata,
        routeChunks=route_chunks,
        routeMarkers=route_markers,
        excursions=excursions,
        excursionRouteChunks=excursion_route_chunks,
        detours=detours,
        detourRouteChunks=detour_route_chunks,
        warnings=lodging_warnings + excursion_warnings + detour_warnings,
    )


def parse_stages(sheet: Worksheet) -> list[Stage]:
    stages: list[Stage] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sequence = to_int(row[0])
        name = clean_text(row[3])
        doc_id = stage_id(sequence, name)
        if sequence is None or name is None or doc_id is None:
            continue
        stages.append(
            Stage(
                id=doc_id,
                sequence=sequence,
                name=name,
                distanceFromPathKm=to_float(row[1]),
                accumulatedDistanceKm=to_float(row[2]),
                segmentLengthKm=to_float(row[4]),
                elevationUpM=to_float(row[5]),
                elevationDownM=to_float(row[6]),
                altitudeM=to_float(row[7]),
                services=StageServices(
                    lodging=to_bool_yn(row[8]),
                    tent=to_bool_yn(row[9]),
                    food=to_bool_yn(row[10]),
                    grocery=to_bool_yn(row[11]),
                    drinkableWater=to_bool_yn(row[12]),
                    nonDrinkableWater=to_bool_yn(row[13]),
                    toilets=to_bool_yn(row[14]),
                    medical=to_bool_yn(row[15]),
                    pharmacy=to_bool_yn(row[16]),
                    atm=to_bool_yn(row[17]),
                    busStop=to_bool_yn(row[18]),
                ),
            )
        )
    return stages


def parse_lodgings(
    sheet: Worksheet,
    stage_lookup: dict[str, str],
    stage_slug_lookup: dict[str, str],
) -> tuple[list[Lodging], list[str]]:
    lodgings: list[Lodging] = []
    warnings: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        stage_sequence = to_int(row[0])
        stage_name = clean_text(row[1])
        lodging_name = clean_text(row[2])
        if stage_sequence is None and stage_name is None and lodging_name is None:
            continue

        doc_id = lodging_id(stage_sequence, stage_name, lodging_name, row_number)
        price_text, price_min, price_max = parse_price(row[3])
        coords = parse_coordinates(row[9])
        resolved_stage_id = resolve_stage_id(stage_name, stage_lookup, stage_slug_lookup)
        if stage_name and not resolved_stage_id:
            warnings.append(f"LODGING row {row_number}: unknown stage '{stage_name}'")

        lodgings.append(
            Lodging(
                id=doc_id,
                stageId=resolved_stage_id,
                stageSequence=stage_sequence,
                stageName=stage_name,
                name=lodging_name,
                minPriceText=price_text,
                priceMinEur=price_min,
                priceMaxEur=price_max,
                distanceFromTrailKm=to_float(row[4]),
                village=clean_text(row[5]),
                type=clean_text(row[6]),
                location=GeoPointValue(latitude=coords[0], longitude=coords[1]) if coords else None,
                address=clean_text(row[10]),
                contact=LodgingContact(
                    phone=clean_text(row[7]),
                    googleMapsUrl=clean_text(row[8]),
                    website=clean_text(row[11]),
                    whatsapp=clean_text(row[12]),
                    email=clean_text(row[13]),
                ),
                openingTime=format_time(row[14]),
                closingTime=format_time(row[15]),
                monthsOpen=clean_text(row[16]),
                capacityPeople=to_int(row[17]),
                checkInTime=format_time(row[18]),
                checkOutTime=format_time(row[19]),
            )
        )

    return lodgings, warnings


def parse_route(
    sheet: Worksheet,
    *,
    stages_by_name: dict[str, str],
    stages_by_slug: dict[str, str],
    chunk_size: int,
    version: int,
) -> tuple[RouteMetadata, list[RouteChunk], list[RouteMarker]]:
    points: list[dict[str, float | str | None]] = []
    markers: list[RouteMarker] = []
    marker_id_counts: dict[str, int] = {}

    for point_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        lat = to_float(row[0])
        lng = to_float(row[1])
        altitude = to_float(row[2])
        distance = to_float(row[3])
        reverse_distance = to_float(row[4])
        stage_name = clean_text(row[5])
        if (
            lat is None
            or lng is None
            or altitude is None
            or distance is None
            or reverse_distance is None
        ):
            continue

        points.append(
            {
                "lat": lat,
                "lng": lng,
                "altitudeM": altitude,
                "distanceKm": distance,
                "reverseDistanceKm": reverse_distance,
            }
        )

        if stage_name:
            resolved_stage_id = resolve_stage_id(stage_name, stages_by_name, stages_by_slug)
            marker_id = resolved_stage_id or f"route-marker-{point_index:05d}"
            marker_id_counts[marker_id] = marker_id_counts.get(marker_id, 0) + 1
            if marker_id_counts[marker_id] > 1:
                marker_id = f"{marker_id}-{point_index:05d}"
            markers.append(
                RouteMarker(
                    id=marker_id,
                    stageId=resolved_stage_id,
                    stageName=stage_name,
                    pointIndex=point_index,
                    distanceKm=distance,
                    reverseDistanceKm=reverse_distance,
                    altitudeM=altitude,
                    location=GeoPointValue(latitude=lat, longitude=lng),
                )
            )

    chunks = build_route_chunks(points, chunk_size)
    lats = [float(point["lat"]) for point in points]
    lngs = [float(point["lng"]) for point in points]
    altitudes = [float(point["altitudeM"]) for point in points]
    distances = [float(point["distanceKm"]) for point in points]

    metadata = RouteMetadata(
        version=version,
        pointCount=len(points),
        chunkCount=len(chunks),
        chunkSize=chunk_size,
        totalDistanceKm=max(distances) if distances else 0,
        minAltitudeM=min(altitudes) if altitudes else 0,
        maxAltitudeM=max(altitudes) if altitudes else 0,
        bounds={
            "minLat": min(lats) if lats else 0,
            "maxLat": max(lats) if lats else 0,
            "minLng": min(lngs) if lngs else 0,
            "maxLng": max(lngs) if lngs else 0,
        },
    )
    return metadata, chunks, markers


def parse_excursions(
    sheet: Worksheet,
    *,
    stages_by_sequence: dict[int, Stage],
    main_route_points: list[list[float]],
    chunk_size: int,
    version: int,
) -> tuple[list[Excursion], dict[str, list[RouteChunk]], list[str]]:
    if len(main_route_points) < 2:
        raise ValueError("Excursions require at least two main E4 route points.")

    records: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[:7]) + [None] * max(0, 7 - len(row))
        lat = to_float(values[0])
        lng = to_float(values[1])
        altitude = to_float(values[2])
        excursion_id_text = clean_text(values[3])
        route_type_text = clean_text(values[4])
        anchor_type_text = clean_text(values[5])
        anchor_stage_sequence = to_int(values[6])

        coordinates_are_blank = lat is None and lng is None and altitude is None
        metadata_is_blank = not any(
            [
                excursion_id_text,
                route_type_text,
                anchor_type_text,
                anchor_stage_sequence is not None,
            ]
        )
        if coordinates_are_blank and metadata_is_blank:
            continue
        if lat is None or lng is None or altitude is None:
            raise ValueError(
                f"Excursions row {row_number} must contain latitude, longitude, and altitude."
            )
        _validate_coordinates(lat, lng, row_number, "Excursions")

        if excursion_id_text:
            if current is not None:
                records.append(current)
            current = {
                "id": slugify(excursion_id_text),
                "routeType": _normalize_route_type(route_type_text, row_number),
                "anchorType": _normalize_anchor_type(anchor_type_text, row_number),
                "anchorStageSequence": anchor_stage_sequence,
                "sourceRow": row_number,
                "points": [],
            }
        elif current is None:
            raise ValueError(
                f"Excursions row {row_number} contains route data before an Excursion ID."
            )
        elif route_type_text or anchor_type_text or anchor_stage_sequence is not None:
            raise ValueError(
                f"Excursions row {row_number} has anchor metadata without an Excursion ID."
            )

        points = current["points"]
        assert isinstance(points, list)
        points.append((lat, lng, altitude))

    if current is not None:
        records.append(current)

    excursions: list[Excursion] = []
    chunks_by_excursion: dict[str, list[RouteChunk]] = {}
    warnings: list[str] = []
    seen_ids: set[str] = set()
    for record in records:
        excursion_id = str(record["id"])
        if excursion_id in seen_ids:
            raise ValueError(f"Duplicate Excursion ID '{excursion_id}'.")
        seen_ids.add(excursion_id)
        excursion, chunks, record_warnings = _build_excursion(
            record,
            stages_by_sequence=stages_by_sequence,
            main_route_points=main_route_points,
            chunk_size=chunk_size,
            version=version,
        )
        excursions.append(excursion)
        chunks_by_excursion[excursion.id] = chunks
        warnings.extend(record_warnings)

    return excursions, chunks_by_excursion, warnings


def _build_excursion(
    record: dict[str, object],
    *,
    stages_by_sequence: dict[int, Stage],
    main_route_points: list[list[float]],
    chunk_size: int,
    version: int,
) -> tuple[Excursion, list[RouteChunk], list[str]]:
    excursion_id = str(record["id"])
    route_type = str(record["routeType"])
    anchor_type = str(record["anchorType"])
    source_row = int(record["sourceRow"])
    source_points = record["points"]
    assert isinstance(source_points, list)
    if len(source_points) < 2:
        raise ValueError(
            f"Excursion '{excursion_id}' starting at row {source_row} needs at least two points."
        )

    typed_points = [(float(point[0]), float(point[1]), float(point[2])) for point in source_points]
    locations = [(point[0], point[1]) for point in typed_points]
    distances = cumulative_distances_km(locations)
    route_distance = distances[-1]
    reverse_distances = [route_distance - distance for distance in distances]
    route_points = [
        {
            "lat": point[0],
            "lng": point[1],
            "altitudeM": point[2],
            "distanceKm": distances[index],
            "reverseDistanceKm": reverse_distances[index],
        }
        for index, point in enumerate(typed_points)
    ]
    chunks = build_route_chunks(route_points, chunk_size)

    altitude_deltas = [
        typed_points[index][2] - typed_points[index - 1][2] for index in range(1, len(typed_points))
    ]
    route_elevation_up = sum(max(delta, 0.0) for delta in altitude_deltas)
    route_elevation_down = sum(max(-delta, 0.0) for delta in altitude_deltas)
    if route_type == "outAndBack":
        total_distance = route_distance * 2
        elevation_up = route_elevation_up + route_elevation_down
        elevation_down = route_elevation_up + route_elevation_down
    else:
        total_distance = route_distance
        elevation_up = route_elevation_up
        elevation_down = route_elevation_down

    anchor_stage_sequence = record["anchorStageSequence"]
    anchor_stage: Stage | None = None
    if anchor_type == "stage":
        if anchor_stage_sequence is None:
            raise ValueError(f"Excursion '{excursion_id}' requires an Anchor Stage.")
        anchor_stage = stages_by_sequence.get(int(anchor_stage_sequence))
        if anchor_stage is None:
            raise ValueError(
                f"Excursion '{excursion_id}' references unknown stage {anchor_stage_sequence}."
            )
    elif anchor_stage_sequence is not None:
        raise ValueError(
            f"Excursion '{excursion_id}' must leave Anchor Stage blank for {anchor_type} anchors."
        )

    trail_locations = [(point[0], point[1]) for point in main_route_points]
    trail_distances = [point[3] for point in main_route_points]
    match_locations = [locations[0]] if anchor_type == "stage" else locations
    match_distances = [0.0] if anchor_type == "stage" else distances
    match = closest_polyline_match(
        match_locations,
        trail_locations,
        first_distances_km=match_distances,
        second_distances_km=trail_distances,
    )

    lats = [point[0] for point in typed_points]
    lngs = [point[1] for point in typed_points]
    altitudes = [point[2] for point in typed_points]
    estimated_minutes = int(total_distance * 12 + elevation_up / 10 + 0.5)
    if estimated_minutes == 0 and (total_distance > 0 or elevation_up > 0):
        estimated_minutes = 1

    warnings: list[str] = []
    if anchor_type in {"stage", "trail"} and match.distanceKm > 1:
        warnings.append(
            f"Excursion '{excursion_id}' is {match.distanceKm:.3f} km from its E4 anchor."
        )
    if route_type == "loop" and _distance_between_locations(locations[0], locations[-1]) > 0.1:
        warnings.append(
            f"Excursion '{excursion_id}' is marked loop but its endpoints differ by >100 m."
        )

    return (
        Excursion(
            id=excursion_id,
            routeType=route_type,
            anchorType=anchor_type,
            anchorStageId=anchor_stage.id if anchor_stage else None,
            anchorStageSequence=anchor_stage.sequence if anchor_stage else None,
            anchorStageName=anchor_stage.name if anchor_stage else None,
            routeVersion=version,
            pointCount=len(typed_points),
            chunkCount=len(chunks),
            chunkSize=chunk_size,
            routeDistanceKm=route_distance,
            totalDistanceKm=total_distance,
            routeElevationUpM=route_elevation_up,
            routeElevationDownM=route_elevation_down,
            elevationUpM=elevation_up,
            elevationDownM=elevation_down,
            estimatedWalkingTimeMinutes=estimated_minutes,
            minAltitudeM=min(altitudes),
            maxAltitudeM=max(altitudes),
            bounds={
                "minLat": min(lats),
                "maxLat": max(lats),
                "minLng": min(lngs),
                "maxLng": max(lngs),
            },
            startLocation=GeoPointValue(latitude=lats[0], longitude=lngs[0]),
            endLocation=GeoPointValue(latitude=lats[-1], longitude=lngs[-1]),
            distanceFromTrailKm=match.distanceKm,
            mainTrailDistanceKm=match.secondDistanceKm,
            excursionConnectionDistanceKm=match.firstDistanceKm,
            excursionConnectionSegmentIndex=match.firstSegmentIndex,
            mainTrailConnectionSegmentIndex=match.secondSegmentIndex,
            connectionLocation=GeoPointValue(
                latitude=match.firstLocation[0], longitude=match.firstLocation[1]
            ),
            trailConnectionLocation=GeoPointValue(
                latitude=match.secondLocation[0], longitude=match.secondLocation[1]
            ),
        ),
        chunks,
        warnings,
    )


def parse_detours(
    sheet: Worksheet,
    *,
    stages: list[Stage],
    main_route_points: list[list[float]],
    chunk_size: int,
    version: int,
) -> tuple[list[Detour], dict[str, list[RouteChunk]], list[str]]:
    if len(main_route_points) < 2:
        raise ValueError("Detours require at least two main E4 route points.")

    records: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[:5]) + [None] * max(0, 5 - len(row))
        lat = to_float(values[0])
        lng = to_float(values[1])
        altitude = to_float(values[2])
        detour_id_text = clean_text(values[3])
        detour_name = clean_text(values[4])

        coordinates_are_blank = lat is None and lng is None and altitude is None
        metadata_is_blank = detour_id_text is None and detour_name is None
        if coordinates_are_blank and metadata_is_blank:
            continue
        if lat is None or lng is None or altitude is None:
            raise ValueError(
                f"Detours row {row_number} must contain latitude, longitude, and altitude."
            )
        _validate_coordinates(lat, lng, row_number, "Detours")

        if detour_id_text:
            if detour_name is None:
                raise ValueError(f"Detours row {row_number} requires a Detour Name.")
            if current is not None:
                records.append(current)
            current = {
                "id": slugify(detour_id_text),
                "name": detour_name,
                "sourceRow": row_number,
                "points": [],
            }
        elif current is None:
            raise ValueError(f"Detours row {row_number} contains route data before a Detour ID.")
        elif detour_name is not None:
            raise ValueError(f"Detours row {row_number} has a Detour Name without a Detour ID.")

        points = current["points"]
        assert isinstance(points, list)
        points.append((lat, lng, altitude))

    if current is not None:
        records.append(current)

    detours: list[Detour] = []
    chunks_by_detour: dict[str, list[RouteChunk]] = {}
    warnings: list[str] = []
    seen_ids: set[str] = set()
    for record in records:
        detour_id = str(record["id"])
        if detour_id in seen_ids:
            raise ValueError(f"Duplicate Detour ID '{detour_id}'.")
        seen_ids.add(detour_id)
        detour, chunks, record_warnings = _build_detour(
            record,
            stages=stages,
            main_route_points=main_route_points,
            chunk_size=chunk_size,
            version=version,
        )
        detours.append(detour)
        chunks_by_detour[detour.id] = chunks
        warnings.extend(record_warnings)

    return detours, chunks_by_detour, warnings


def _build_detour(
    record: dict[str, object],
    *,
    stages: list[Stage],
    main_route_points: list[list[float]],
    chunk_size: int,
    version: int,
) -> tuple[Detour, list[RouteChunk], list[str]]:
    detour_id = str(record["id"])
    detour_name = str(record["name"])
    source_row = int(record["sourceRow"])
    source_points = record["points"]
    assert isinstance(source_points, list)
    if len(source_points) < 2:
        raise ValueError(
            f"Detour '{detour_id}' starting at row {source_row} needs at least two points."
        )

    typed_points = [(float(point[0]), float(point[1]), float(point[2])) for point in source_points]
    locations = [(point[0], point[1]) for point in typed_points]
    distances = cumulative_distances_km(locations)
    route_distance = distances[-1]
    reverse_distances = [route_distance - distance for distance in distances]
    route_points = [
        {
            "lat": point[0],
            "lng": point[1],
            "altitudeM": point[2],
            "distanceKm": distances[index],
            "reverseDistanceKm": reverse_distances[index],
        }
        for index, point in enumerate(typed_points)
    ]
    chunks = build_route_chunks(route_points, chunk_size)
    elevation_up, elevation_down = _elevation_totals(typed_points)

    trail_locations = [(point[0], point[1]) for point in main_route_points]
    trail_distances = [point[3] for point in main_route_points]
    start_match = closest_polyline_match(
        [locations[0]],
        trail_locations,
        first_distances_km=[0.0],
        second_distances_km=trail_distances,
    )
    end_match = closest_polyline_match(
        [locations[-1]],
        trail_locations,
        first_distances_km=[route_distance],
        second_distances_km=trail_distances,
    )
    replaced_points = _main_route_section(main_route_points, start_match, end_match)
    replaced_elevation_up, replaced_elevation_down = _elevation_totals(
        [(point[0], point[1], point[2]) for point in replaced_points]
    )
    replaced_distance = abs(end_match.secondDistanceKm - start_match.secondDistanceKm)
    estimated_minutes = _estimated_walking_minutes(route_distance, elevation_up)
    replaced_estimated_minutes = _estimated_walking_minutes(
        replaced_distance,
        replaced_elevation_up,
    )
    trail_separations = point_distances_to_polyline_km(locations, trail_locations)
    affected_stages = _affected_detour_stages(
        stages,
        start_match.secondDistanceKm,
        end_match.secondDistanceKm,
    )

    warnings: list[str] = []
    for label, match in (("start", start_match), ("end", end_match)):
        if match.distanceKm > 0.1:
            warnings.append(
                f"Detour '{detour_id}' {label} is {match.distanceKm:.3f} km from the E4."
            )
    if replaced_distance < 0.01:
        warnings.append(
            f"Detour '{detour_id}' start and end connect to nearly the same E4 location."
        )

    lats = [point[0] for point in typed_points]
    lngs = [point[1] for point in typed_points]
    altitudes = [point[2] for point in typed_points]
    return (
        Detour(
            id=detour_id,
            name=detour_name,
            routeVersion=version,
            pointCount=len(typed_points),
            chunkCount=len(chunks),
            chunkSize=chunk_size,
            routeDistanceKm=route_distance,
            elevationUpM=elevation_up,
            elevationDownM=elevation_down,
            estimatedWalkingTimeMinutes=estimated_minutes,
            replacedMainTrailDistanceKm=replaced_distance,
            replacedElevationUpM=replaced_elevation_up,
            replacedElevationDownM=replaced_elevation_down,
            replacedEstimatedWalkingTimeMinutes=replaced_estimated_minutes,
            distanceDifferenceKm=route_distance - replaced_distance,
            elevationUpDifferenceM=elevation_up - replaced_elevation_up,
            elevationDownDifferenceM=elevation_down - replaced_elevation_down,
            estimatedWalkingTimeDifferenceMinutes=estimated_minutes - replaced_estimated_minutes,
            averageDistanceFromTrailKm=(
                sum(trail_separations) / len(trail_separations) if trail_separations else 0.0
            ),
            maximumDistanceFromTrailKm=max(trail_separations) if trail_separations else 0.0,
            minAltitudeM=min(altitudes),
            maxAltitudeM=max(altitudes),
            bounds={
                "minLat": min(lats),
                "maxLat": max(lats),
                "minLng": min(lngs),
                "maxLng": max(lngs),
            },
            startConnection=_detour_connection(start_match, locations[0]),
            endConnection=_detour_connection(end_match, locations[-1]),
            affectedStageIds=[stage.id for stage in affected_stages],
            affectedStageSequences=[stage.sequence for stage in affected_stages],
            affectedStageNames=[stage.name for stage in affected_stages],
        ),
        chunks,
        warnings,
    )


def _detour_connection(
    match: PolylineMatch,
    route_location: tuple[float, float],
) -> DetourConnection:
    return DetourConnection(
        distanceFromTrailKm=match.distanceKm,
        mainTrailDistanceKm=match.secondDistanceKm,
        mainTrailSegmentIndex=match.secondSegmentIndex,
        routeLocation=GeoPointValue(
            latitude=route_location[0],
            longitude=route_location[1],
        ),
        trailLocation=GeoPointValue(
            latitude=match.secondLocation[0],
            longitude=match.secondLocation[1],
        ),
    )


def _main_route_section(
    points: list[list[float]],
    start_match: PolylineMatch,
    end_match: PolylineMatch,
) -> list[list[float]]:
    start_point = _interpolate_main_route_point(
        points,
        start_match.secondSegmentIndex,
        start_match.secondFraction,
    )
    end_point = _interpolate_main_route_point(
        points,
        end_match.secondSegmentIndex,
        end_match.secondFraction,
    )
    start_index = start_match.secondSegmentIndex
    end_index = end_match.secondSegmentIndex
    if start_match.secondDistanceKm <= end_match.secondDistanceKm:
        middle = points[start_index + 1 : end_index + 1]
    else:
        middle = list(reversed(points[end_index + 1 : start_index + 1]))
    return [start_point, *middle, end_point]


def _interpolate_main_route_point(
    points: list[list[float]],
    segment_index: int,
    fraction: float,
) -> list[float]:
    start = points[segment_index]
    end = points[min(segment_index + 1, len(points) - 1)]
    return [
        start[value_index] + (end[value_index] - start[value_index]) * fraction
        for value_index in range(5)
    ]


def _elevation_totals(
    points: list[tuple[float, float, float]],
) -> tuple[float, float]:
    deltas = [points[index][2] - points[index - 1][2] for index in range(1, len(points))]
    return (
        sum(max(delta, 0.0) for delta in deltas),
        sum(max(-delta, 0.0) for delta in deltas),
    )


def _estimated_walking_minutes(distance_km: float, elevation_up_m: float) -> int:
    minutes = int(distance_km * 12 + elevation_up_m / 10 + 0.5)
    if minutes == 0 and (distance_km > 0 or elevation_up_m > 0):
        return 1
    return minutes


def _affected_detour_stages(
    stages: list[Stage],
    start_distance_km: float,
    end_distance_km: float,
) -> list[Stage]:
    located_stages = sorted(
        (stage for stage in stages if stage.accumulatedDistanceKm is not None),
        key=lambda stage: float(stage.accumulatedDistanceKm or 0),
    )
    minimum = min(start_distance_km, end_distance_km)
    maximum = max(start_distance_km, end_distance_km)
    affected: list[Stage] = []
    for index in range(1, len(located_stages)):
        previous_distance = float(located_stages[index - 1].accumulatedDistanceKm or 0)
        stage_distance = float(located_stages[index].accumulatedDistanceKm or 0)
        leg_minimum = min(previous_distance, stage_distance)
        leg_maximum = max(previous_distance, stage_distance)
        if leg_maximum > minimum and leg_minimum < maximum:
            affected.append(located_stages[index])
    return affected


def _normalize_route_type(value: str | None, row_number: int) -> str:
    normalized = slugify(value).replace("-", "") if value else ""
    route_types = {
        "oneway": "oneWay",
        "outandback": "outAndBack",
        "loop": "loop",
    }
    if normalized not in route_types:
        raise ValueError(
            f"Excursions row {row_number} Route Type must be oneWay, outAndBack, or loop."
        )
    return route_types[normalized]


def _normalize_anchor_type(value: str | None, row_number: int) -> str:
    normalized = (value or "").strip().lower()
    if normalized not in {"stage", "trail", "standalone"}:
        raise ValueError(
            f"Excursions row {row_number} Anchor Type must be stage, trail, or standalone."
        )
    return normalized


def _validate_coordinates(
    lat: float,
    lng: float,
    row_number: int,
    sheet_name: str,
) -> None:
    if not -90 <= lat <= 90 or not -180 <= lng <= 180:
        raise ValueError(f"{sheet_name} row {row_number} contains invalid GPS coordinates.")


def _distance_between_locations(
    first: tuple[float, float],
    second: tuple[float, float],
) -> float:
    return cumulative_distances_km([first, second])[-1]


def resolve_stage_id(
    stage_name: str | None,
    stages_by_name: dict[str, str],
    stages_by_slug: dict[str, str],
) -> str | None:
    if not stage_name:
        return None
    return stages_by_name.get(stage_name) or stages_by_slug.get(slugify(stage_name))


def build_route_chunks(
    points: list[dict[str, float | str | None]], chunk_size: int
) -> list[RouteChunk]:
    chunks: list[RouteChunk] = []
    for chunk_index, start in enumerate(range(0, len(points), chunk_size)):
        chunk_points = points[start : start + chunk_size]
        end = start + len(chunk_points) - 1
        compact_points = [
            value
            for point in chunk_points
            for value in [
                float(point["lat"]),
                float(point["lng"]),
                float(point["altitudeM"]),
                float(point["distanceKm"]),
                float(point["reverseDistanceKm"]),
            ]
        ]
        grouped_points = group_flat_points(compact_points)
        lats = [point[0] for point in grouped_points]
        lngs = [point[1] for point in grouped_points]
        distances = [point[3] for point in grouped_points]
        chunks.append(
            RouteChunk(
                id=f"{chunk_index:04d}",
                chunkIndex=chunk_index,
                startPointIndex=start,
                endPointIndex=end,
                startDistanceKm=min(distances),
                endDistanceKm=max(distances),
                bounds={
                    "minLat": min(lats),
                    "maxLat": max(lats),
                    "minLng": min(lngs),
                    "maxLng": max(lngs),
                },
                points=compact_points,
            )
        )
    return chunks


def group_flat_points(points: list[float], stride: int = 5) -> list[list[float]]:
    return [points[index : index + stride] for index in range(0, len(points), stride)]


def flatten_route_points(chunks: list[RouteChunk]) -> list[list[float]]:
    points: list[list[float]] = []
    for chunk in sorted(chunks, key=lambda item: item.chunkIndex):
        points.extend(group_flat_points(chunk.points))
    return points
